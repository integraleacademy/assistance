import datetime
from io import BytesIO

from openpyxl import load_workbook

import app as application
from crm_exports import CRM_EXPORT_HEADERS


def _client(tmp_path, monkeypatch, authenticated=True):
    monkeypatch.setattr(application, "DATA_FILE", str(tmp_path / "data.json"))
    application.app.config.update(TESTING=True, SESSION_COOKIE_SECURE=False)
    client = application.app.test_client()
    if authenticated:
        with client.session_transaction() as session:
            session["user_email"] = "clement@integraleacademy.com"
    return client


def _save_contacts(contacts):
    data = application.load_data()
    data["crm_contacts"] = contacts
    application.save_data(data)


def _worksheet(response):
    workbook = load_workbook(BytesIO(response.data))
    return workbook.active


def test_exports_page_is_private_and_available_without_sidebar_entry(tmp_path, monkeypatch):
    anonymous = _client(tmp_path, monkeypatch, authenticated=False)
    assert anonymous.get("/crm/exports").status_code == 302
    assert anonymous.get("/api/crm/exports/aps").status_code == 302

    client = _client(tmp_path, monkeypatch)
    page = client.get("/crm/exports")
    crm_js = open(
        application.app.root_path + "/static/crm.js", encoding="utf-8"
    ).read()
    template = open(
        application.app.root_path + "/templates/crm.html", encoding="utf-8"
    ).read()

    assert page.status_code == 200
    assert b'"section": "exports"' in page.data
    assert 'id="dashboardExportsButton"' in crm_js
    assert "function exportsPage()" in crm_js
    assert 'data-nav="exports"' not in template


def test_excel_export_has_requested_columns_and_only_converted_contacts(
        tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    _save_contacts([
        {
            "id": "converted-aps",
            "statut": "Converti",
            "formation": "APS",
            "converted_at": "2026-08-20T08:30:00+00:00",
            "nom": "MARTIN",
            "prenom": "Lina",
            "mail": "lina@example.com",
            "telephone": "+33601020304",
            "gclid": "",
            "formulaire": {"gclid": "gclid-aps-123"},
        },
        {
            "id": "lead-aps",
            "statut": "Nouveaux",
            "formation": "APS",
            "nom": "PASINSCRIT",
        },
        {
            "id": "converted-a3p",
            "statut": "Converti",
            "formation": "A3P",
            "nom": "AUTREFORMATION",
        },
    ])

    response = client.get("/api/crm/exports/aps")
    worksheet = _worksheet(response)

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["Content-Disposition"]
    assert "fichier_aps_" in response.headers["Content-Disposition"]
    assert response.headers["Cache-Control"] == "private, no-store, max-age=0"
    assert tuple(cell.value for cell in worksheet[1]) == CRM_EXPORT_HEADERS
    assert worksheet.max_row == 2
    assert tuple(cell.value for cell in worksheet[2]) == (
        "APS",
        datetime.datetime(2026, 8, 20, 10, 30),
        "MARTIN",
        "Lina",
        "lina@example.com",
        "+33601020304",
        "gclid-aps-123",
    )
    assert worksheet["F2"].data_type == "s"
    assert worksheet["G2"].data_type == "s"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:G2"


def test_each_formation_has_its_own_file_and_desp_paths_are_separated(
        tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    contacts = [
        {"id": "a3p", "statut": "Converti", "formation": "A3P", "nom": "A3P"},
        {"id": "aps", "statut": "Converti", "formation": "APS", "nom": "APS"},
        {
            "id": "desp-initial", "statut": "Converti", "formation": "DESP",
            "desp_type": "INITIAL", "nom": "INITIAL",
        },
        {
            "id": "desp-vae", "statut": "Converti", "formation": "DESP",
            "desp_type": "VAE", "nom": "VAE",
        },
        {
            "id": "ssiap", "statut": "Converti", "formation": "SSIAP 1",
            "nom": "SSIAP",
        },
        {
            "id": "vtc", "statut": "Converti", "formation": "Chauffeur VTC",
            "nom": "VTC",
        },
    ]
    _save_contacts(contacts)

    expectations = {
        "a3p": ("A3P", "A3P"),
        "aps": ("APS", "APS"),
        "desp-initial": ("DESP initial", "INITIAL"),
        "desp-vae": ("DESP VAE", "VAE"),
        "ssiap": ("SSIAP 1", "SSIAP"),
        "chauffeur-vtc": ("Chauffeur VTC", "VTC"),
    }
    for export_key, (formation_label, last_name) in expectations.items():
        response = client.get(f"/api/crm/exports/{export_key}")
        worksheet = _worksheet(response)
        assert response.status_code == 200
        assert worksheet.max_row == 2
        assert worksheet["A2"].value == formation_label
        assert worksheet["C2"].value == last_name


def test_unknown_export_is_not_downloadable(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    assert client.get("/api/crm/exports/inconnue").status_code == 404
