"""Construction des exports Excel des inscriptions du CRM."""

from __future__ import annotations

import datetime
import re
import unicodedata
from io import BytesIO

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CRM_EXPORT_DEFINITIONS = {
    "a3p": {
        "label": "A3P",
        "filename": "fichier_a3p",
        "sheet": "A3P",
    },
    "aps": {
        "label": "APS",
        "filename": "fichier_aps",
        "sheet": "APS",
    },
    "desp-initial": {
        "label": "DESP initial",
        "filename": "fichier_desp_initial",
        "sheet": "DESP initial",
    },
    "desp-vae": {
        "label": "DESP VAE",
        "filename": "fichier_desp_vae",
        "sheet": "DESP VAE",
    },
    "ssiap": {
        "label": "SSIAP 1",
        "filename": "fichier_ssiap",
        "sheet": "SSIAP 1",
    },
    "chauffeur-vtc": {
        "label": "Chauffeur VTC",
        "filename": "fichier_chauffeur_vtc",
        "sheet": "Chauffeur VTC",
    },
}

CRM_EXPORT_HEADERS = (
    "Nom de la formation",
    "Date de conversion / inscription",
    "Nom",
    "Prénom",
    "Mail",
    "Téléphone",
    "GCLID (si disponible)",
)

_PARIS_TIMEZONE = pytz.timezone("Europe/Paris")


def _normalized(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text
                   if not unicodedata.combining(character))
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def crm_export_key_for_contact(contact):
    """Retourne le fichier auquel appartient une fiche CRM."""
    formation = _normalized(contact.get("formation"))
    desp_type = _normalized(contact.get("desp_type"))
    combined = f"{formation} {desp_type}".strip()

    if "DESP" in formation or "DIRIGEANT" in formation:
        return "desp-vae" if "VAE" in combined else "desp-initial"
    if "A3P" in formation:
        return "a3p"
    if "APS" in formation.split():
        return "aps"
    if "SSIAP" in formation:
        return "ssiap"
    if "VTC" in formation:
        return "chauffeur-vtc"
    return None


def _conversion_source_value(contact):
    """Retrouve la date la plus fiable, y compris pour les anciennes fiches."""
    if contact.get("converted_at"):
        return contact["converted_at"]

    activities = []
    for activity in contact.get("activities") or []:
        if not isinstance(activity, dict) or not activity.get("date"):
            continue
        kind = _normalized(activity.get("kind"))
        title = _normalized(activity.get("title"))
        if kind == "CONVERSION" or (kind == "STATUT" and "CONVERTI" in title):
            activities.append(activity["date"])
    if activities:
        return max(activities, key=lambda value: str(value))

    return (
        contact.get("status_changed_at")
        or contact.get("updated_at")
        or contact.get("created_at")
        or ""
    )


def _excel_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
        return datetime.datetime.combine(value, datetime.time.min)
    if isinstance(value, datetime.datetime):
        parsed = value
    else:
        raw = str(value).strip()
        try:
            parsed = datetime.datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return raw
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(_PARIS_TIMEZONE).replace(tzinfo=None)
    return parsed


def _gclid(contact):
    value = str(contact.get("gclid") or "").strip()
    if value:
        return value
    form = contact.get("formulaire")
    if isinstance(form, dict):
        value = str(form.get("gclid") or "").strip()
        if value:
            return value
    if _normalized(contact.get("google_ads_identifier_type")) == "GCLID":
        return str(contact.get("google_ads_identifier") or "").strip()
    return ""


def _write_text(cell, value):
    """Force le texte afin de préserver les 0, les + et de bloquer les formules."""
    cell.value = str(value or "")
    cell.data_type = "s"
    cell.number_format = "@"


def _export_rows(contacts, export_key):
    definition = CRM_EXPORT_DEFINITIONS[export_key]
    rows = []
    for contact in contacts or []:
        if not isinstance(contact, dict):
            continue
        if _normalized(contact.get("statut")) != "CONVERTI":
            continue
        if crm_export_key_for_contact(contact) != export_key:
            continue
        converted_at = _excel_datetime(_conversion_source_value(contact))
        rows.append((
            definition["label"],
            converted_at,
            contact.get("nom", ""),
            contact.get("prenom", ""),
            contact.get("mail", ""),
            contact.get("telephone", ""),
            _gclid(contact),
        ))

    def sort_key(row):
        date_value = row[1]
        if isinstance(date_value, datetime.datetime):
            timestamp = date_value.timestamp()
        else:
            timestamp = float("-inf")
        return (-timestamp, _normalized(row[2]), _normalized(row[3]))

    return sorted(rows, key=sort_key)


def build_crm_export_workbook(contacts, export_key):
    """Construit un classeur XLSX en mémoire pour une formation."""
    if export_key not in CRM_EXPORT_DEFINITIONS:
        raise KeyError(export_key)

    definition = CRM_EXPORT_DEFINITIONS[export_key]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = definition["sheet"]
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="173E8D")
    header_font = Font(color="FFFFFF", bold=True)
    for column, label in enumerate(CRM_EXPORT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 26

    for row_number, row in enumerate(_export_rows(contacts, export_key), start=2):
        for column in (1, 3, 4, 5, 6, 7):
            _write_text(worksheet.cell(row=row_number, column=column), row[column - 1])
        date_cell = worksheet.cell(row=row_number, column=2, value=row[1])
        if isinstance(row[1], datetime.datetime):
            date_cell.number_format = "dd/mm/yyyy hh:mm"
        date_cell.alignment = Alignment(vertical="top")
        if row_number % 2 == 0:
            for cell in worksheet[row_number]:
                cell.fill = PatternFill("solid", fgColor="F5F8FE")

    widths = (23, 31, 22, 22, 34, 20, 48)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.auto_filter.ref = f"A1:G{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    workbook.properties.creator = "Intégrale Academy"
    workbook.properties.title = f"Inscrits - {definition['label']}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def crm_export_filename(export_key, today=None):
    definition = CRM_EXPORT_DEFINITIONS[export_key]
    export_date = today or datetime.datetime.now(_PARIS_TIMEZONE).date()
    return f"{definition['filename']}_{export_date.isoformat()}.xlsx"
